from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd

from src.config import settings


@dataclass(frozen=True)
class PreflightReport:
    input_path: Path
    template_pf: Path
    template_pj: Path
    output_root: Path
    rows: int
    invoices_total: int
    invoices_pf: int
    invoices_pj: int


def _require(cond: bool, msg: str) -> None:
    if not cond:
        raise ValueError(msg)


def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def preflight_checks(df: pd.DataFrame) -> PreflightReport:
    """
    Valida ambiente/arquivos/config/dados ANTES do processamento.
    Lança exceções com mensagens claras se algo estiver fora do esperado.
    Retorna um relatório com contagens para você logar/mostrar.
    """
    input_path = Path(settings.input_file)
    template_pf = Path(settings.template_pf)
    template_pj = Path(settings.template_pj)
    output_root = Path(settings.output_dir)

    # ===== Checks de arquivos/pastas =====
    _require(input_path.exists(), f"Arquivo de entrada não encontrado: {input_path.resolve()}")
    _require(template_pf.exists(), f"Template PF não encontrado: {template_pf.resolve()}")
    _require(template_pj.exists(), f"Template PJ não encontrado: {template_pj.resolve()}")

    # Garante pasta de saída
    output_root.mkdir(parents=True, exist_ok=True)

    # ===== Checks de DataFrame =====
    _require(len(df) > 0, "Arquivo de entrada não possui linhas para processar.")

    df = _normalize_cols(df)

    # Colunas mínimas para o fluxo atual (PF/PJ + transações)
    required_cols: Iterable[str] = [
        settings.group_by_column.lower(),          # documento_cliente
        settings.client_type_column.lower(),       # tipo_cliente
        "nome_cliente",
        settings.month_ref_column.lower(),         # mes_fatura
        settings.card_number_column.lower(),       # numero_cartao
        "estabelecimento",
        "valor_compra",
        "qtd_parcelas",
        "valor_parcela",
    ]

    missing = [c for c in required_cols if c not in df.columns]
    _require(not missing, f"Colunas obrigatórias ausentes: {missing}")

    # ===== Qualidade dos dados =====
    # tipo_cliente só PF/PJ
    invalid_types = (
        df[settings.client_type_column.lower()]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .unique()
        .tolist()
    )
    allowed = {"PF", "PJ"}
    bad = [t for t in invalid_types if t and t not in allowed]
    _require(not bad, f"Valores inválidos em '{settings.client_type_column}': {bad}. Aceitos: {sorted(allowed)}")

    # Documento não vazio
    doc_col = settings.group_by_column.lower()
    empty_docs = df[doc_col].fillna("").astype(str).str.strip().eq("").sum()
    _require(empty_docs == 0, f"Existem {empty_docs} linhas sem '{settings.group_by_column}' preenchido.")

    # Converte números e valida
    def to_float_series(col: str) -> pd.Series:
        return pd.to_numeric(df[col].astype(str).str.replace(",", "."), errors="coerce")

    valor_parcela = to_float_series("valor_parcela")
    qtd_parcelas = pd.to_numeric(df["qtd_parcelas"].astype(str).str.replace(",", "."), errors="coerce")

    _require(valor_parcela.notna().all(), "Há valores inválidos na coluna 'valor_parcela' (não numéricos).")
    _require(qtd_parcelas.notna().all(), "Há valores inválidos na coluna 'qtd_parcelas' (não numéricos).")
    _require((qtd_parcelas >= 1).all(), "Há 'qtd_parcelas' menor que 1.")
    _require((valor_parcela >= 0).all(), "Há 'valor_parcela' negativo.")

    # ===== Check de template: aba existe =====
    from openpyxl import load_workbook

    for tpath, label in [(template_pf, "PF"), (template_pj, "PJ")]:
        wb = load_workbook(tpath)
        _require(
            settings.sheet_template in wb.sheetnames,
            f"Template {label} '{tpath.name}' não contém a aba '{settings.sheet_template}'. Abas: {wb.sheetnames}",
        )

    # ===== Contagens (quantas faturas serão geradas) =====
    # Uma fatura por documento_cliente (do jeito que o group_invoices faz hoje)
    df_type = df[settings.client_type_column.lower()].astype(str).str.strip().str.upper()
    df_doc = df[doc_col].astype(str).str.strip()

    unique_docs = df_doc.nunique(dropna=True)
    pf_docs = df.loc[df_type == "PF", doc_col].astype(str).str.strip().nunique(dropna=True)
    pj_docs = df.loc[df_type == "PJ", doc_col].astype(str).str.strip().nunique(dropna=True)

    # Check de “explosão” (proteção simples)
    _require(unique_docs <= 5000, f"Número muito alto de faturas ({unique_docs}). Verifique agrupamento/arquivo.")

    return PreflightReport(
        input_path=input_path,
        template_pf=template_pf,
        template_pj=template_pj,
        output_root=output_root,
        rows=int(len(df)),
        invoices_total=int(unique_docs),
        invoices_pf=int(pf_docs),
        invoices_pj=int(pj_docs),
    )
