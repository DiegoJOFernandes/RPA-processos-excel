from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from src.config import settings


def _safe_set_cell(ws: Worksheet, cell_addr: str, value) -> None:
    """
    Escreve valor em célula, tratando casos em que a célula está dentro de um range mesclado.
    Se estiver mesclada, escreve na célula top-left do merge.
    """
    # Se a célula estiver em algum merge, encontra o range e grava no top-left
    for merged_range in ws.merged_cells.ranges:
        if cell_addr in merged_range:
            ws[merged_range.start_cell.coordinate].value = value
            return

    # Não é mesclada: grava normalmente
    ws[cell_addr].value = value


def fill_invoice_template(
    header: dict,
    items: list[dict],
    template_file: Path,
    output_path: Path
) -> None:
    # Abre o template específico (PF ou PJ) escolhido no main.py
    wb = load_workbook(template_file)

    # ✅ Validação clara da aba (evita KeyError confuso)
    if settings.sheet_template not in wb.sheetnames:
        raise KeyError(
            f"Aba '{settings.sheet_template}' não encontrada no template "
            f"'{template_file.name}'. Abas disponíveis: {wb.sheetnames}"
        )

    ws = wb[settings.sheet_template]

    # Cabeçalho principal
    _safe_set_cell(ws, settings.cell_doc, header["documento"])
    _safe_set_cell(ws, settings.cell_name, header["nome"])
    _safe_set_cell(ws, settings.cell_date, header["data_emissao"])
    _safe_set_cell(ws, settings.cell_total, header["total"])

    # ✅ Campos extras (cartão / mês)
    _safe_set_cell(ws, settings.cell_month_ref, header.get("mes_referencia", ""))
    _safe_set_cell(ws, settings.cell_card_number, header.get("numero_cartao", ""))
    _safe_set_cell(ws, settings.cell_monthly_sum, header.get("total_mensal", header["total"]))

    # Itens (limpa área antes)
    start = settings.items_start_row
    max_items = settings.max_items

    for i in range(max_items):
        r = start + i
        _safe_set_cell(ws, f"{settings.col_item_desc}{r}", None)
        _safe_set_cell(ws, f"{settings.col_item_qty}{r}", None)
        _safe_set_cell(ws, f"{settings.col_item_unit}{r}", None)
        _safe_set_cell(ws, f"{settings.col_item_total}{r}", None)

    # Preenche itens até o limite
    for i, item in enumerate(items[:max_items]):
        r = start + i
        _safe_set_cell(ws, f"{settings.col_item_desc}{r}", item["descricao"])
        _safe_set_cell(ws, f"{settings.col_item_qty}{r}", item["quantidade"])
        _safe_set_cell(ws, f"{settings.col_item_unit}{r}", item["valor_unitario"])
        _safe_set_cell(ws, f"{settings.col_item_total}{r}", item["valor_total"])

    # Garante pasta de saída e salva o arquivo final
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
